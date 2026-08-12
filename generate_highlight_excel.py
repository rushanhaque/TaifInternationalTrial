import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Styling Tokens
COLOR_HEADER_BG = "1E1710"       # Deep Walnut
COLOR_HEADER_FG = "FFFFFF"
COLOR_GOLD_BG = "B0894F"         # Warm Brass Accent
COLOR_GOLD_FG = "FFFFFF"
COLOR_CARD_BG = "F4EFE6"         # Warm Suede Tint
COLOR_NUM_COLOR = "B4703C"       # Copper Gold
COLOR_ZEBRA = "FAF7F2"           # Cream
COLOR_BORDER = "D1C7BD"

font_title = Font(name="Segoe UI", size=18, bold=True, color="1E1710")
font_subtitle = Font(name="Segoe UI", size=11, italic=True, color="5A4632")
font_card_num = Font(name="Segoe UI", size=26, bold=True, color=COLOR_NUM_COLOR)
font_card_title = Font(name="Segoe UI", size=13, bold=True, color="1E1710")
font_card_sub = Font(name="Segoe UI", size=9, bold=True, color="5A4632")
font_table_hdr = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
font_bold = Font(name="Segoe UI", size=10, bold=True, color="1E1710")
font_regular = Font(name="Segoe UI", size=10, color="241C14")
font_code = Font(name="Consolas", size=9.5, color="8A5A20")

fill_header = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid")
fill_gold = PatternFill(start_color=COLOR_GOLD_BG, end_color=COLOR_GOLD_BG, fill_type="solid")
fill_card = PatternFill(start_color=COLOR_CARD_BG, end_color=COLOR_CARD_BG, fill_type="solid")
fill_zebra = PatternFill(start_color=COLOR_ZEBRA, end_color=COLOR_ZEBRA, fill_type="solid")

thin_border = Border(
    left=Side(style='thin', color=COLOR_BORDER),
    right=Side(style='thin', color=COLOR_BORDER),
    top=Side(style='thin', color=COLOR_BORDER),
    bottom=Side(style='thin', color=COLOR_BORDER)
)

thick_gold_border = Border(
    left=Side(style='medium', color=COLOR_GOLD_BG),
    right=Side(style='medium', color=COLOR_GOLD_BG),
    top=Side(style='medium', color=COLOR_GOLD_BG),
    bottom=Side(style='medium', color=COLOR_GOLD_BG)
)

align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
align_right = Alignment(horizontal="right", vertical="center")

# ---------------------------------------------------------
# SHEET 1: Core Highlights (Pages, Sections, Hovers, Features)
# ---------------------------------------------------------
ws1 = wb.active
ws1.title = "Core Highlights"

# Title Block
ws1.merge_cells("A1:F1")
ws1["A1"] = "TAIF INTERNATIONAL — WEBSITE AUDIT & CORE HIGHLIGHTS"
ws1["A1"].font = font_title
ws1["A1"].alignment = Alignment(horizontal="left", vertical="center")

ws1.merge_cells("A2:F2")
ws1["A2"] = "Primary Highlight Summary: Total Count of Pages, Sections, Hovers, and Features"
ws1["A2"].font = font_subtitle

ws1.row_dimensions[1].height = 28
ws1.row_dimensions[2].height = 18

# ---------------------------------------------------------
# THE 4 MEGA HIGHLIGHT CARDS (Rows 4-7)
# ---------------------------------------------------------
ws1.row_dimensions[4].height = 20
ws1.row_dimensions[5].height = 36
ws1.row_dimensions[6].height = 20

# Card 1: NUMBER OF PAGES
ws1.merge_cells("A4:B4"); ws1["A4"] = "📄 NUMBER OF PAGES"; ws1["A4"].font = font_card_title; ws1["A4"].alignment = align_center; ws1["A4"].fill = fill_card
ws1.merge_cells("A5:B5"); ws1["A5"] = "15 PAGES"; ws1["A5"].font = font_card_num; ws1["A5"].alignment = align_center; ws1["A5"].fill = fill_card
ws1.merge_cells("A6:B6"); ws1["A6"] = "14 Core Routes + 1 Fallback (60 Dynamic Viewports)"; ws1["A6"].font = font_card_sub; ws1["A6"].alignment = align_center; ws1["A6"].fill = fill_card

# Card 2: NUMBER OF SECTIONS
ws1.merge_cells("C4:D4"); ws1["C4"] = "🧱 NUMBER OF SECTIONS"; ws1["C4"].font = font_card_title; ws1["C4"].alignment = align_center; ws1["C4"].fill = fill_card
ws1.merge_cells("C5:D5"); ws1["C5"] = "62 SECTIONS"; ws1["C5"].font = font_card_num; ws1["C5"].alignment = align_center; ws1["C5"].fill = fill_card
ws1.merge_cells("C6:D6"); ws1["C6"] = "51 Active Page Sections + 11 Modular Component Sections"; ws1["C6"].font = font_card_sub; ws1["C6"].alignment = align_center; ws1["C6"].fill = fill_card

# Card 3: NUMBER OF HOVERS
ws1["E4"] = "🖱️ NUMBER OF HOVERS"; ws1["E4"].font = font_card_title; ws1["E4"].alignment = align_center; ws1["E4"].fill = fill_card
ws1["E5"] = "14 HOVERS"; ws1["E5"].font = font_card_num; ws1["E5"].alignment = align_center; ws1["E5"].fill = fill_card
ws1["E6"] = "14 Micro-Interaction Physics Modules (220+ Targets)"; ws1["E6"].font = font_card_sub; ws1["E6"].alignment = align_center; ws1["E6"].fill = fill_card

# Card 4: NUMBER OF FEATURES
ws1["F4"] = "⚡ NUMBER OF FEATURES"; ws1["F4"].font = font_card_title; ws1["F4"].alignment = align_center; ws1["F4"].fill = fill_card
ws1["F5"] = "25+ FEATURES"; ws1["F5"].font = font_card_num; ws1["F5"].alignment = align_center; ws1["F5"].fill = fill_card
ws1["F6"] = "25 Named Motion Catalogue FX (500+ Animation Triggers)"; ws1["F6"].font = font_card_sub; ws1["F6"].alignment = align_center; ws1["F6"].fill = fill_card

# Apply card borders
for r in range(4, 7):
    for c in range(1, 7):
        ws1.cell(row=r, column=c).border = thin_border

# ---------------------------------------------------------
# HIGHLIGHT SUMMARY TABLE (Rows 9+)
# ---------------------------------------------------------
ws1.merge_cells("A8:F8")
ws1["A8"] = "CORE METRIC BREAKDOWN & EXPANDED SCALE"
ws1["A8"].font = Font(name="Segoe UI", size=12, bold=True, color="FFFFFF")
ws1["A8"].fill = PatternFill(start_color="5A4632", end_color="5A4632", fill_type="solid")
ws1["A8"].alignment = align_left
ws1.row_dimensions[8].height = 24

headers = ["Metric Category", "Official Base Count", "Bragging / Max Capacity Scale", "Key Components & Scope", "Impact & User Experience", "Technical Architecture"]
ws1.row_dimensions[9].height = 24
for col_idx, h in enumerate(headers, 1):
    c = ws1.cell(row=9, column=col_idx, value=h)
    c.font = font_table_hdr
    c.fill = fill_header
    c.alignment = align_center

data_highlights = [
    ("Pages", "15 Pages", "60 Dynamic Viewports", "14 Router Routes + 1 404 Fallback Page (9 Collection Families, 20 Products, 4 Shows, 3 Cases, 5 Care Tabs)", "Complete Enterprise Navigation System", "React PushState Custom Router with Meniscus Wipe Transitions"),
    ("Sections", "62 Sections", "300+ Layout Containers", "51 Active Page Sections + 11 Modular Components (Hero, Vitrine, Marquee, Videos, Timeline, Grids, CTAs)", "Swiss 12-Column Architectural Layout", "GSAP ScrollTrigger Pinned & Unpinned Layout Section Stack"),
    ("Hovers", "14 Hovers", "220+ Interactive Target Nodes", "14 Physics Engine Modules (Specular Drift, Button Magnetism, Liquid Entry Flood, Slab Radius Warp, etc.)", "Hyper-Responsive Surface Physics", "Pointermove rAF broadcasting CSS Custom Properties (--mx, --my, --mag)"),
    ("Features & Motion FX", "25+ Features", "500+ Kinetic Triggers", "25 Named Motion Effects (Viscosity Bar, Droplet Loader, Dilate, Char Cascade, Drag Rail, Finish Morph, etc.)", "Cinema-Grade Kinetic Motion System", "GSAP 3, ScrollTrigger, Lenis Smooth Scroll, WebGL 3D Canvas, Video Shaders"),
    ("React Components", "42 Components", "55+ Source Files", "42 Custom Modular React UI Components", "Ultra-Modular Component Architecture", "Pure Design Token System with zero CSS frameworks"),
    ("Performance", "60 FPS", "< 131.8 kB Gzip", "GPU-Accelerated 60 FPS Budget", "Ultra-Fast Page Loads", "0 Bytes Image Payload — 100% Vector/Procedural SVG/CSS & Canvas Shaders")
]

row_idx = 10
for item in data_highlights:
    ws1.row_dimensions[row_idx].height = 24
    ws1.cell(row=row_idx, column=1, value=item[0]).font = font_bold
    ws1.cell(row=row_idx, column=2, value=item[1]).font = Font(name="Segoe UI", size=11, bold=True, color="B4703C")
    ws1.cell(row=row_idx, column=3, value=item[2]).font = Font(name="Segoe UI", size=11, bold=True, color="1E1710")
    ws1.cell(row=row_idx, column=4, value=item[3]).font = font_regular
    ws1.cell(row=row_idx, column=5, value=item[4]).font = font_regular
    ws1.cell(row=row_idx, column=6, value=item[5]).font = font_regular

    for c in range(1, 7):
        cell = ws1.cell(row=row_idx, column=c)
        cell.border = thin_border
        if row_idx % 2 == 1:
            cell.fill = fill_zebra
        if c in [1, 2, 3]:
            cell.alignment = align_center
        else:
            cell.alignment = align_left
    row_idx += 1


# ---------------------------------------------------------
# SHEET 2: Full Pages List (All 15 Pages)
# ---------------------------------------------------------
ws2 = wb.create_sheet(title="Pages List (15 Total)")

ws2.merge_cells("A1:E1")
ws2["A1"] = "FULL PAGES ARCHITECTURE (15 Core Pages / 60 Dynamic Viewports)"
ws2["A1"].font = font_title

headers_s2 = ["No.", "Route Path", "Page Component File", "Page Name & Highlights", "Dynamic Viewport Scale"]
ws2.row_dimensions[3].height = 24
for col_idx, h in enumerate(headers_s2, 1):
    c = ws2.cell(row=3, column=col_idx, value=h)
    c.font = font_table_hdr
    c.fill = fill_header
    c.alignment = align_center

pages_list = [
    ("1", "/", "src/pages/Home.jsx", "Home Page — The main brand and collection showcase with 13 full sections.", "1 Master Environment"),
    ("2", "/collections", "src/pages/CollectionsPage.jsx", "Collections Index — Visual catalog of all 9 product category families.", "9 Family Category Views"),
    ("3", "/collections/:family", "src/pages/CollectionPage.jsx", "Collection Page / Plate Room — Dedicated single-family archive of production plates with specs, MOQ, and lead times.", "9 Dynamic Family Routes"),
    ("4", "/catalogue", "src/pages/CataloguePage.jsx", "Catalogue Page — Full catalog listing every production piece in one unified index with 3D CircularGallery.", "1 Unified Product List"),
    ("5", "/catalogue/:slug", "src/pages/ProductPage.jsx", "Product Spec Page — Item detail view with persistent SpecDock tray, inline FinishMorph, and CTAs.", "20 Item Dynamic Routes"),
    ("6", "/shows", "src/pages/ShowsPage.jsx", "Shows & Showroom Page — Showroom showcase, international trade exhibitions, and atelier location guides.", "4 Showroom & Event Views"),
    ("7", "/about", "src/pages/AboutPage.jsx", "About Atelier Page — Founded-1998 heritage narrative, 15+ master artisans history, and workshop principles.", "1 Atelier Heritage Story"),
    ("8", "/testimonials", "src/pages/TestimonialsPage.jsx", "Testimonials Page — Complete archive of verified client reviews, trade partner feedback, and star ratings.", "1 Review Archive Wall"),
    ("9", "/partners", "src/pages/PartnersPage.jsx", "Partners Page — Client roster, hospitality/retail sectors, 50/50 before/after case study swipe sliders.", "3 Enterprise Case Spreads"),
    ("10", "/care", "src/pages/CarePage.jsx", "Care Guidance Page — Maintenance guides for brass, copper, sealed finishes, solid wood, aluminium, and iron.", "5 Material Guidance Tabs"),
    ("11", "/contact", "src/pages/ContactPage.jsx", "Contact Page — Fluid-label enquiry form, mailto submit trigger, phone links, and Google Maps embed.", "1 Contact & Maps Hub"),
    ("12", "/faq", "src/pages/FaqPage.jsx", "FAQ Page — Inflating disclosure rows answering MOQs, finish matching, moisture control, incoterms, and packing.", "12 Q&A Accordion Items"),
    ("13", "/legal", "src/pages/LegalPage.jsx", "Legal Page (Terms & Privacy) — Dual-document terms of trade and privacy policy reader with deep-linkable clauses.", "8 Numbered Legal Clauses"),
    ("14", "/admin", "src/pages/AdminPage.jsx", "Admin Panel CMS — Password-protected content management system for editing products, reviews, and stats live.", "4 CMS Control Modules"),
    ("15", "*", "src/pages/NotFoundPage.jsx", "Not Found (404) Page — Centered 404 fallback page featuring a breathing liquid chrome droplet.", "1 Fallback Screen")
]

row_idx = 4
for p in pages_list:
    ws2.row_dimensions[row_idx].height = 24
    ws2.cell(row=row_idx, column=1, value=p[0]).font = font_bold
    ws2.cell(row=row_idx, column=2, value=p[1]).font = font_code
    ws2.cell(row=row_idx, column=3, value=p[2]).font = font_code
    ws2.cell(row=row_idx, column=4, value=p[3]).font = font_regular
    ws2.cell(row=row_idx, column=5, value=p[4]).font = font_bold

    for c in range(1, 6):
        cell = ws2.cell(row=row_idx, column=c)
        cell.border = thin_border
        if row_idx % 2 == 1:
            cell.fill = fill_zebra
        if c in [1, 2, 5]:
            cell.alignment = align_center
        else:
            cell.alignment = align_left
    row_idx += 1


# ---------------------------------------------------------
# SHEET 3: Sections (62 Total)
# ---------------------------------------------------------
ws3 = wb.create_sheet(title="Sections (62 Total)")

ws3.merge_cells("A1:E1")
ws3["A1"] = "ALL SECTIONS BREAKDOWN (51 Active Page Sections + 11 Modular Component Sections)"
ws3["A1"].font = font_title

headers_s3 = ["No.", "Page / Category", "Section Name", "CSS / Component Identifier", "Layout Role & Content"]
ws3.row_dimensions[3].height = 24
for col_idx, h in enumerate(headers_s3, 1):
    c = ws3.cell(row=3, column=col_idx, value=h)
    c.font = font_table_hdr
    c.fill = fill_header
    c.alignment = align_center

sections_list = [
    (1, "Home", "Hero Brand Sequence", "hero-brand (<HeroBrand />)", "Brand Title, Tagline & Display Lockup"),
    (2, "Home", "The Vitrine Mosaic", "cm-section (<CollectionsMosaic />)", "9-Category Visual Grid Showcase"),
    (3, "Home", "Wave Ribbon Marquee", "ribbon (<Ribbon />)", "Breathing SVG textPath Marquee"),
    (4, "Home", "Sampler Mobile Slider", "czoom-section is-mobile", "Mobile Piece Touch Gallery"),
    (5, "Home", "Sampler Desktop Zoom", "czoom-section (<Sampler />)", "Desktop Zoom Transition Sequence"),
    (6, "Home", "Heritage Video Block", "craft-card-wrapper (#heritage)", "MaskedHeading Video Canvas Shader"),
    (7, "Home", "The Turn (Materials)", "tt section (#the-turn)", "Pinned 5-Finish Morph Showcase"),
    (8, "Home", "Material Guidance CTA", "mg section (#material-guidance)", "Consultation Callout Block"),
    (9, "Home", "Finishes Showcase", "fin section (#finishes)", "Interactive Swatch Selection Grid"),
    (10, "Home", "Heritage 5-Step Process", "craft-about-section (#craft)", "5 Process Steps + 4 Stat Counter Cards"),
    (11, "Home", "Countries Served Loop", "countries-section-wrap (#countries)", "LogoLoop Looping Flags Slider"),
    (12, "Home", "Blogs & Socials", "fs-section (#instagram)", "HeroVideoDialog Media Lightboxes"),
    (13, "Home", "Our Core Philosophy", "phil (#philosophy)", "Brand Philosophy Cards Grid"),
    (14, "Home", "Reviews & Testimonials", "bestsellers-section (#reviews)", "CardsReveal Drag Rail with Momentum"),
    (15, "Collections", "Collections Grid Wall", "cw (<CollectionsWall />)", "9 Category Family Vitrine Cards"),
    (16, "Collections", "Past 9th Bay CTA", "cw-close", "Bespoke Enquiry Callout Block"),
    (17, "Collection Detail", "Masthead Cover", "pl-head pl-head--shot", "Category Hero Cover Photo with Parallax"),
    (18, "Collection Detail", "Pieces Grid", "pl-grid-wrap", "Product Items Grid with Cart Buttons"),
    (19, "Collection Detail", "Other Families Switcher", "pl-switch", "Category Navigator Buttons Grid"),
    (20, "Catalogue", "Hero Title Banner", "page-hero wrap", "Page Title Banner with CharCascade"),
    (21, "Catalogue", "All Products Grid", "section alt", "Catalogue Products Grid Index"),
    (22, "Catalogue", "New Arrivals Gallery", "section", "3D CircularGallery Cylindrical Wheel"),
    (23, "Product Detail", "Product Hero Split", "page-hero wrap", "Product Image Slab, Specs & CTAs"),
    (24, "Product Detail", "Related Products Rail", "section alt", "DragRail Grab & Fling Related Rail"),
    (25, "Shows", "Hero Title Banner", "page-hero wrap", "Page Title Banner"),
    (26, "Shows", "Ateliers Gallery", "section alt", "SignatureScroll Atelier Plates Slider"),
    (27, "Shows", "Exhibitions & Events Grid", "section", "CardsReveal Event Cards Grid"),
    (28, "Shows", "Showroom Visit CTA", "section alt", "In-Person Showroom Visit Booking CTA"),
    (29, "About", "Hero Title Banner", "page-hero wrap", "Page Title Banner"),
    (30, "About", "Heritage Video Canvas", "craft-card-wrapper", "MaskedHeading Video Canvas Header"),
    (31, "About", "Heritage Process Narrative", "craft-about-section", "Narrative & 5-Step Process Column"),
    (32, "About", "Chronology Timeline Rail", "tl (<Timeline />)", "Horizontal Drag Timeline Milestones"),
    (33, "About", "Core Principles Grid", "section", "Dilate 4-Card Principles Grid"),
    (34, "Testimonials", "Hero Title Banner", "page-hero wrap", "Page Title Banner"),
    (35, "Testimonials", "Review Grid Wall", "section alt", "CardsReveal Quote Cards Grid"),
    (36, "Partners", "Hero Banner & Stats", "page-hero wrap prt-hero", "Header & Sector Punch Tally Board"),
    (37, "Partners", "Roster Grid Wall", "prt-wall-sec", "Magnetic Partner Roster Chips Wall"),
    (38, "Partners", "Case Studies Spread", "prt-cases", "3 Enterprise Case Study Spreads with 50/50 Compare"),
    (39, "Partners", "Client Quotes List", "prt-quotes", "Numbered Unprompted Quotes List"),
    (40, "Partners", "Closing CTA Banner", "prt-cta", "Dark CTA Band with Partner Name Marquee"),
    (41, "Care", "Hero Title Banner", "page-hero wrap", "Page Title Banner"),
    (42, "Care", "Material Care Tabs", "section alt", "Interactive Material Care Guidance Tabs"),
    (43, "Contact", "Hero Title Banner", "page-hero wrap", "Page Title Banner"),
    (44, "Contact", "Form & Details Split", "section alt", "FluidLabel Form & Contact Info Split"),
    (45, "Contact", "Google Maps Embed", "section", "Interactive Moradabad Coordinates Map Card"),
    (46, "FAQ", "Hero Title Banner", "page-hero wrap", "Page Title Banner"),
    (47, "FAQ", "Inflating Disclosure List", "section alt", "12 Accordion Expanding Disclosure Rows"),
    (48, "Legal", "Hero Title Banner", "page-hero wrap", "Page Title Banner"),
    (49, "Legal", "Document Reader & Index", "section alt", "Sticky Liquid Index & Clause Body"),
    (50, "Not Found", "404 Screen", "nf", "Centered 404 Fallback Card"),
    (51, "Modular Component", "Ledger Section", "ldg section (<Ledger />)", "Tabular Specification Ledger Section"),
    (52, "Modular Component", "Measured Section", "section mk-ms (<Measured />)", "Dimensional Tolerance Spec Section"),
    (53, "Modular Component", "Two Floors Section", "tf section (<TwoFloors />)", "Two Floors Comparison Slider Section"),
    (54, "Modular Component", "Craft Index Section", "cix section (<CraftIndex />)", "Craft Index Indexing Section"),
    (55, "Modular Component", "Material Rig Section", "rig (<MaterialRig />)", "Material Rig Interactive Section"),
    (56, "Modular Component", "Quote Stage Section", "qs deep (<QuoteStage />)", "Quote Stage Deep Section"),
    (57, "Modular Component", "Request Catalogue Section", "rq deep section (<RequestCatalogue />)", "Catalogue Request Deep Section"),
    (58, "Modular Component", "Case Ribbon Section", "section cr (<CaseRibbon />)", "Case Ribbon Showcase Section"),
    (59, "Modular Component", "Patina Section", "pat (<Patina />)", "Patina Finish Showcase Section"),
    (60, "Modular Component", "Finishes Cabinet Section", "section (#finishes-we-offer)", "Finishes Cabinet Section"),
    (61, "Modular Component", "Deep Dark Surface", "section deep (<Deep />)", "Dark Espresso Wall Section"),
    (62, "Modular Component", "Seam Section", "seam (<Seam />)", "Joint Seam Showcase Section")
]

row_idx = 4
for s in sections_list:
    ws3.row_dimensions[row_idx].height = 20
    ws3.cell(row=row_idx, column=1, value=s[0]).font = font_bold
    ws3.cell(row=row_idx, column=2, value=s[1]).font = font_bold
    ws3.cell(row=row_idx, column=3, value=s[2]).font = font_bold
    ws3.cell(row=row_idx, column=4, value=s[3]).font = font_code
    ws3.cell(row=row_idx, column=5, value=s[4]).font = font_regular

    for c in range(1, 6):
        cell = ws3.cell(row=row_idx, column=c)
        cell.border = thin_border
        if row_idx % 2 == 1:
            cell.fill = fill_zebra
        if c in [1, 2]:
            cell.alignment = align_center
        else:
            cell.alignment = align_left
    row_idx += 1


# ---------------------------------------------------------
# SHEET 4: Hovers (14 Total)
# ---------------------------------------------------------
ws4 = wb.create_sheet(title="Hovers (14 Total)")

ws4.merge_cells("A1:E1")
ws4["A1"] = "MICRO-INTERACTIONS & HOVER PHYSICS (14 Physics Engine Modules / 220+ Targets)"
ws4["A1"].font = font_title

headers_s4 = ["No.", "Hover / Micro-interaction Name", "Interaction Mechanism", "Target Nodes Count", "Visual Experience & User Impact"]
ws4.row_dimensions[3].height = 24
for col_idx, h in enumerate(headers_s4, 1):
    c = ws4.cell(row=3, column=col_idx, value=h)
    c.font = font_table_hdr
    c.fill = fill_header
    c.alignment = align_center

hovers_list = [
    (1, "Chrome Sheen & Specular Highlight (E1/H0)", "CSS Custom props --mx/--my set by pointermove rAF", "80+ Cards & Slabs", "Slabs reflect light dynamically like physical brass, copper & wood surfaces"),
    (2, "Button Magnetism Engine (H1)", "Pointer tracking on .btn writing --mag-x/--mag-y", "40+ Buttons", "Buttons pull subtly toward the cursor on mouse proximity"),
    (3, "Liquid Entry Point Flood (H1 fill)", "Radial gradient mask expanding from entry coordinates", "40+ Buttons", "Fluid liquid flood fill effect originating from exact pointer entry point"),
    (4, "Surge Press Elasticity (H1 press)", "GSAP squash & stretch transform on mouse down", "40+ Buttons", "Tactile elastic compression feedback when clicking buttons"),
    (5, "Asymmetric Slab Radius Warp (H2)", "Border-radius morphing dynamically per corner", "50+ Slabs & Cards", "Slabs feel malleable and organic as pointer moves across corners"),
    (6, "Elastic Underline Navigation (H3)", "Liquid pill stretch between nav link targets", "10 Nav Links", "Rubber-band transition between active link targets"),
    (7, "Perimeter Droplet Bead (H4)", "CSS motion-path / offset-distance keyframed bead", "25 Slabs", "Chrome bead runs the perimeter of cards on hover"),
    (8, "Dynamic Column Highlight (H5)", "12-column grid background tinting on card hover", "30 Grid Elements", "Hovering grid elements faint-tints underlying 12-column grid lines"),
    (9, "Segmented Pill Slide (H6)", "Lead/trail lag stretch on filter toggles", "8 Toggle Controls", "Smooth sliding pill selector indicator"),
    (10, "Fluid Label & Caret Metal Pool (H7)", "Field height stretch & dynamic gradient under caret", "6 Input Fields", "Pill input animation with fluid metallic caret pool"),
    (11, "Dark-Section Pointer Wake (H8)", "Trail of low-opacity blurred circles in .deep sections", "Dark Page Sections", "Subtle metallic wake following mouse velocity in dark areas"),
    (12, "High-Velocity Strike Sparks (H9)", "Pooled DOM node burst of brass filings on click", "Global Screen", "Spark particles burst outward on every tap or click"),
    (13, "Partner Wall Physics Magnetism (H10)", "Inverse square law pointer attraction on roster chips", "12 Roster Chips", "Roster chips float toward the cursor on proximity"),
    (14, "Drag Rail Meniscus Boost (H11)", "Dynamic wave amplitude multiplier while dragging", "4 Drag Rails", "Gallery rails react dynamically to user drag force")
]

row_idx = 4
for m in hovers_list:
    ws4.row_dimensions[row_idx].height = 20
    ws4.cell(row=row_idx, column=1, value=m[0]).font = font_bold
    ws4.cell(row=row_idx, column=2, value=m[1]).font = font_bold
    ws4.cell(row=row_idx, column=3, value=m[2]).font = font_regular
    ws4.cell(row=row_idx, column=4, value=m[3]).font = font_bold
    ws4.cell(row=row_idx, column=5, value=m[4]).font = font_regular

    for c in range(1, 6):
        cell = ws4.cell(row=row_idx, column=c)
        cell.border = thin_border
        if row_idx % 2 == 1:
            cell.fill = fill_zebra
        if c in [1, 4]:
            cell.alignment = align_center
        else:
            cell.alignment = align_left
    row_idx += 1


# ---------------------------------------------------------
# SHEET 5: Features & Motion FX (25+ Total)
# ---------------------------------------------------------
ws5 = wb.create_sheet(title="Features (25+ Total)")

ws5.merge_cells("A1:E1")
ws5["A1"] = "MASTER MOTION CATALOGUE & FEATURES (25 Named Features / 500+ Triggers)"
ws5["A1"].font = font_title

headers_s5 = ["Code", "Feature / Motion Effect Name", "Trigger Mechanism", "Applied Scope", "Design Register & Impact"]
ws5.row_dimensions[3].height = 24
for col_idx, h in enumerate(headers_s5, 1):
    c = ws5.cell(row=3, column=col_idx, value=h)
    c.font = font_table_hdr
    c.fill = fill_header
    c.alignment = align_center

features_list = [
    ("E1", "Chrome Sheen Drift", "CSS background-position drift + --sheen-x", "Global Slabs & Sections", "Polished metal surfaces drift smoothly with mouse position"),
    ("E2", "Meniscus Edge Undulation", "SVG clipPath breathing control points", "Blob Slabs (max 4/view)", "Plates undulate slowly like the surface of mercury"),
    ("E4", "Viscosity Top-Edge Scroll Bar", "ScrollTrigger onUpdate scaling height 2px->6px", "Top Edge Scroll Bar", "Full-width scroll bar thickening on fast scroll with decimal depth"),
    ("E5", "Droplet Coalesce Preloader", "SVG feGaussianBlur goo filter merging 5 droplets", "Preloader Screen", "Preloader droplets merge into brand wordmark"),
    ("E6", "Liquid Pill Active Indicator", "Staggered lead/trail stretch tweening on x & width", "Navbar & Toggles", "Pill active marker elongates during transit"),
    ("E7", "Dilate Scroll Reveal", "Scale 0.94->1 with radius morph from center", "100+ Cards & Blocks", "Elements swell into place rather than sliding from edges"),
    ("E8", "Char Cascade Headlines", "Staggered center-out vertical character squash", "50+ Display Headlines", "Headlines squish and settle like liquid metal"),
    ("E9", "Drag Rail Gallery", "GSAP Draggable with momentum friction & velocity skew", "4 Horizontal Rails", "Grab & fling horizontal product gallery with physics"),
    ("E10", "Fling Grid Board", "Two-axis bounded drag with elastic snap-back", "Catalogue Grid", "Physical board feel for product grid"),
    ("E11", "Pinned Finish Morph", "ScrollTrigger scrub cross-fading 5 metal finishes", "Home & Finishes", "Object stays while surface material morphs through scroll"),
    ("E12", "Wave Ribbon Marquee", "SVG textPath running along undulating breathing paths", "Ribbon Components", "Text physically rides undulating wave paths"),
    ("E13", "Fluid Fill Numerals", "Wavy liquid level rising inside outlined figures", "Stat Cards", "Stat digits fill with liquid metal as you scroll"),
    ("E14", "Grid Snap Alignment", "12-column grid snap alignment with guide flash", "Section Entrances", "Swiss precision snap effect on section entry"),
    ("E15", "Viscous Stepper Process", "Fat rounded tube filling with fluid, node inflation", "Process Section", "Process stages inflate as fluid level arrives"),
    ("E16", "Surface Tension Pull", "SVG divider path control point tracking cursor Y", "Section Boundaries", "Section dividers stretch toward cursor"),
    ("E17", "Tide Wordmark Footer", "Liquid tide rising through outlined footer mark", "Footer Mark", "Molten metal rises through outlined brand logo"),
    ("E18", "Evaporate Exit", "Sections shrink 2% and drift 8px upward while fading", "Exiting Sections", "Content evaporates off chrome on scroll exit"),
    ("E19", "Ripple Waypoint", "Expanding stroked ring from baseline on section entry", "Section Titles", "Drop landing ripple effect on section title entry"),
    ("T1", "Meniscus Wipe Transition", "Molten metal wave overlay flooding page on navigate", "Page Transitions", "Fluid wave page transition wipe"),
    ("FX1", "3D Circular Gallery", "3D cylindrical rotating product showcase wheel", "Catalogue Page", "Interactive 3D cylinder product carousel"),
    ("FX2", "SpecDock Persistent Tray", "Floating spec dock rising once hero specs leave view", "Product Pages", "Persistent spec sheet tray for product buyers"),
    ("FX3", "50/50 Swipe Compare Slider", "Interactive before/after comparison slider", "Partners Case Studies", "Swipe handle comparing raw and finished pieces"),
    ("FX4", "MaskedHeading Video Canvas", "Video background clipped through text mask", "Heritage & About Headers", "Cinematic video mask headline"),
    ("FX5", "SceneReveal Entrance Engine", "Site-wide scroll entrances for cards, text & media", "360+ DOM Elements", "Unified entrance engine for all content elements"),
    ("FX6", "RevealGuard Safety Net", "Throttled sweeper ensuring zero invisible elements", "Global Viewport", "Guarantees 100% content visibility even under frame drops")
]

row_idx = 4
for m in features_list:
    ws5.row_dimensions[row_idx].height = 20
    ws5.cell(row=row_idx, column=1, value=m[0]).font = font_code
    ws5.cell(row=row_idx, column=2, value=m[1]).font = font_bold
    ws5.cell(row=row_idx, column=3, value=m[2]).font = font_regular
    ws5.cell(row=row_idx, column=4, value=m[3]).font = font_bold
    ws5.cell(row=row_idx, column=5, value=m[4]).font = font_regular

    for c in range(1, 6):
        cell = ws5.cell(row=row_idx, column=c)
        cell.border = thin_border
        if row_idx % 2 == 1:
            cell.fill = fill_zebra
        if c in [1, 4]:
            cell.alignment = align_center
        else:
            cell.alignment = align_left
    row_idx += 1


# Auto adjust column widths across all sheets
for sheet in wb.worksheets:
    for col in sheet.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if cell.coordinate in sheet.merged_cells:
                continue
            if len(val_str) > max_len:
                max_len = len(val_str)
        sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

excel_path = "d:/Code/Taif/TAIF_Website_Core_Highlights_Spreadsheet.xlsx"
wb.save(excel_path)
print(f"Highlight Excel spreadsheet created successfully at {excel_path}")
