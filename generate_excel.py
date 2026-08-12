import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Colors
COLOR_HEADER_BG = "241C14"       # Deep Walnut/Graphite
COLOR_HEADER_FG = "FFFFFF"
COLOR_ACCENT_BG = "B0894F"       # Warm Brass Accent
COLOR_ACCENT_FG = "FFFFFF"
COLOR_ZEBRA = "FDFBF7"           # Alabaster / Light Cream
COLOR_CARD_BG = "F4EFE6"         # Warm Suede/Bronze tint
COLOR_BORDER = "D1C7BD"

font_title = Font(name="Segoe UI", size=16, bold=True, color="241C14")
font_subtitle = Font(name="Segoe UI", size=10, italic=True, color="5A4632")
font_section = Font(name="Segoe UI", size=12, bold=True, color=COLOR_HEADER_BG)
font_header = Font(name="Segoe UI", size=11, bold=True, color=COLOR_HEADER_FG)
font_card_num = Font(name="Segoe UI", size=20, bold=True, color="B0894F")
font_card_lbl = Font(name="Segoe UI", size=9, bold=True, color="5A4632")
font_bold = Font(name="Segoe UI", size=10, bold=True)
font_regular = Font(name="Segoe UI", size=10)

fill_header = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid")
fill_accent = PatternFill(start_color=COLOR_ACCENT_BG, end_color=COLOR_ACCENT_BG, fill_type="solid")
fill_zebra = PatternFill(start_color=COLOR_ZEBRA, end_color=COLOR_ZEBRA, fill_type="solid")
fill_card = PatternFill(start_color=COLOR_CARD_BG, end_color=COLOR_CARD_BG, fill_type="solid")

thin_border = Border(
    left=Side(style='thin', color=COLOR_BORDER),
    right=Side(style='thin', color=COLOR_BORDER),
    top=Side(style='thin', color=COLOR_BORDER),
    bottom=Side(style='thin', color=COLOR_BORDER)
)

align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
align_right = Alignment(horizontal="right", vertical="center")

# ---------------------------------------------------------
# SHEET 1: Executive Summary
# ---------------------------------------------------------
ws1 = wb.active
ws1.title = "Executive Summary"

ws1.merge_cells("A1:G1")
ws1["A1"] = "TAIF INTERNATIONAL — WEBSITE ARCHITECTURE & AUDIT METRICS"
ws1["A1"].font = font_title
ws1["A1"].alignment = Alignment(horizontal="left", vertical="center")

ws1.merge_cells("A2:G2")
ws1["A2"] = "Comprehensive Enterprise Performance Audit & Motion System Quantification"
ws1["A2"].font = font_subtitle

# KPI Cards Block (Rows 4-6)
kpis = [
    ("40+", "Dynamic Page Views & States", "A4:B5"),
    ("62", "Engineered Layout Sections", "C4:D5"),
    ("14", "Advanced Micro-Interactions", "E4:E5"),
    ("25+", "Named Motion FX Engine", "F4:F5"),
    ("120+", "Kinetic Scroll Triggers", "G4:G5")
]

ws1.row_dimensions[4].height = 28
ws1.row_dimensions[5].height = 18

ws1["A4"] = "40+"
ws1["A5"] = "Dynamic Page Views & States"
ws1.merge_cells("A4:B4")
ws1.merge_cells("A5:B5")

ws1["C4"] = "62"
ws1["C5"] = "Engineered Layout Sections"
ws1.merge_cells("C4:D4")
ws1.merge_cells("C5:D5")

ws1["E4"] = "14"
ws1["E5"] = "Micro-Interactions & Hovers"

ws1["F4"] = "25+"
ws1["F5"] = "Named Motion FX Engine"

ws1["G4"] = "120+"
ws1["G5"] = "Kinetic Scroll Triggers"

for col_idx in range(1, 8):
    col_letter = get_column_letter(col_idx)
    cell_top = ws1[f"{col_letter}4"]
    cell_bot = ws1[f"{col_letter}5"]
    cell_top.font = font_card_num
    cell_top.alignment = align_center
    cell_top.fill = fill_card
    cell_bot.font = font_card_lbl
    cell_bot.alignment = align_center
    cell_bot.fill = fill_card

# High Level Overview Table
ws1["A8"] = "Metric Category"
ws1["B8"] = "Standard Metric Count"
ws1["C8"] = "Expanded / Dynamic Metric Count"
ws1["D8"] = "Impact & Complexity Rating"
ws1["E8"] = "Engineering Highlight"
ws1.merge_cells("E8:G8")

ws1.row_dimensions[8].height = 24
for col in range(1, 8):
    c = ws1.cell(row=8, column=col)
    c.font = font_header
    c.fill = fill_header
    c.alignment = align_center

summary_data = [
    ("Core Application Routes & Pages", 15, "40+ Dynamic Views (9 Families + 20 Products + 4 Shows)", "Enterprise Multi-page", "Custom PushState router with Meniscus Wipe transition"),
    ("Layout Sections & Viewports", 51, "62 Sections (51 Active + 11 Modular Components)", "High Visual Depth", "Swiss 12-column kinetic grid layouts"),
    ("Micro-Interactions & Hover FX", 14, "14 Engine Physics Features", "Hyper-Responsive", "Pointer-tracking specular highlights & magnetic buttons"),
    ("Motion Engine & Animations", 25, "120+ Keyframed & Scroll Triggers", "Cinema-Grade Motion", "GSAP 3, ScrollTrigger, Lenis smooth scroll & canvas UI"),
    ("Modular Component Library", 42, "55+ JS/JSX Source Modules", "Modular & Scalable", "Reusable design tokens with zero image payload"),
    ("Performance & Optimization", "60 FPS", "< 131.8 kB gzipped bundle", "Ultra Performance", "Zero image payload with GPU-accelerated CSS/SVG rendering")
]

row_idx = 9
for item in summary_data:
    ws1.row_dimensions[row_idx].height = 22
    ws1.cell(row=row_idx, column=1, value=item[0]).font = font_bold
    ws1.cell(row=row_idx, column=2, value=item[1]).font = font_regular
    ws1.cell(row=row_idx, column=3, value=item[2]).font = font_regular
    ws1.cell(row=row_idx, column=4, value=item[3]).font = font_bold
    
    ws1.merge_cells(start_row=row_idx, start_column=5, end_row=row_idx, end_column=7)
    ws1.cell(row=row_idx, column=5, value=item[4]).font = font_regular

    for c in range(1, 8):
        cell = ws1.cell(row=row_idx, column=c)
        cell.border = thin_border
        if row_idx % 2 == 1:
            cell.fill = fill_zebra
        if c in [2, 3, 4]:
            cell.alignment = align_center
        else:
            cell.alignment = align_left

    row_idx += 1


# ---------------------------------------------------------
# SHEET 2: Pages & Dynamic Views
# ---------------------------------------------------------
ws2 = wb.create_sheet(title="Pages & Views")

ws2.merge_cells("A1:E1")
ws2["A1"] = "PAGE & ROUTING ARCHITECTURE (15 Core Routes / 40+ Dynamic Views)"
ws2["A1"].font = font_title

headers_s2 = ["Route Path", "Page Component", "Page Name / Description", "Dynamic View State Count", "Key Interactive Features"]
ws2.row_dimensions[3].height = 24
for col_idx, h in enumerate(headers_s2, 1):
    c = ws2.cell(row=3, column=col_idx, value=h)
    c.font = font_header
    c.fill = fill_header
    c.alignment = align_center

pages_data = [
    ("/", "Home.jsx", "Home — Hero, Vitrine, Craft, Finishes & Philosophy", "1 Master Hub", "HeroBrand, CollectionsMosaic, FinishMorph, DragRail, Sampler"),
    ("/collections", "CollectionsPage.jsx", "Collections Index — 9 Category Families Grid", "9 Category Views", "CollectionsWall, 9-family preview plates, Bespoke CTA"),
    ("/collections/:family", "CollectionPage.jsx", "Collection Detail — Specific Category View", "9 Dynamic Routes", "Velora landscape cards, spec rails, category switcher"),
    ("/catalogue", "CataloguePage.jsx", "Full Catalogue — Complete Product Index", "1 Master List", "CircularGallery 3D wheel, product grid filter"),
    ("/catalogue/:slug", "ProductPage.jsx", "Product Spec Page — Item Detail View", "20 Item Routes", "SpecDock persistent tray, inline FinishMorph, Specular CTAs"),
    ("/shows", "ShowsPage.jsx", "Shows & Showroom — Exhibitions & Atelier", "4 Show Views", "SignatureScroll slider, event cards, visit booking form"),
    ("/about", "AboutPage.jsx", "About Atelier — 1998 Heritage & Principles", "1 Atelier Story", "MaskedHeading video background, Timeline horizontal rail"),
    ("/testimonials", "TestimonialsPage.jsx", "Testimonials — Trade Reviews & Feedback", "1 Review Wall", "CardsReveal grid, star ratings, category filters"),
    ("/partners", "PartnersPage.jsx", "Partners — Roster & Case Studies", "3 Case Studies", "Interactive before/after swipe slider, magnetic partner wall"),
    ("/care", "CarePage.jsx", "Care Guidance — Material Maintenance Tabs", "5 Material Tabs", "Brass, Copper, Solid Wood, Aluminium, Iron care guides"),
    ("/contact", "ContactPage.jsx", "Contact — Enquiry Form & Maps", "1 Contact Hub", "Fluid label fields, interactive Google Maps coordinate embed"),
    ("/faq", "FaqPage.jsx", "FAQ — Inflating Disclosure List", "12 Q&A Items", "4 categories, expanding disclosure pill rows"),
    ("/legal", "LegalPage.jsx", "Terms & Privacy — Legal Document Reader", "8 Clauses", "Sticky liquid pill index, paragraph deep-link copy engine"),
    ("/admin", "AdminPage.jsx", "Admin Panel — Content Management System", "4 CMS Modules", "Passcode gate, live product/review editor, publish engine"),
    ("*", "NotFoundPage.jsx", "404 Not Found — Fallback Page", "1 Fallback View", "Breathing droplet animation, return CTA")
]

row_idx = 4
for p in pages_data:
    ws2.row_dimensions[row_idx].height = 20
    ws2.cell(row=row_idx, column=1, value=p[0]).font = font_bold
    ws2.cell(row=row_idx, column=2, value=p[1]).font = font_regular
    ws2.cell(row=row_idx, column=3, value=p[2]).font = font_regular
    ws2.cell(row=row_idx, column=4, value=p[3]).font = font_bold
    ws2.cell(row=row_idx, column=5, value=p[4]).font = font_regular

    for c in range(1, 6):
        cell = ws2.cell(row=row_idx, column=c)
        cell.border = thin_border
        if row_idx % 2 == 1:
            cell.fill = fill_zebra
        if c in [1, 4]:
            cell.alignment = align_center
        else:
            cell.alignment = align_left
    row_idx += 1


# ---------------------------------------------------------
# SHEET 3: Sections Audit
# ---------------------------------------------------------
ws3 = wb.create_sheet(title="Sections Audit")

ws3.merge_cells("A1:E1")
ws3["A1"] = "COMPLETE SECTIONS AUDIT (51 Active Page Sections + 11 Modular Components)"
ws3["A1"].font = font_title

headers_s3 = ["Page / Category", "Active Section Name", "CSS / Component Identifier", "Content Type & Role", "Interactive Features"]
ws3.row_dimensions[3].height = 24
for col_idx, h in enumerate(headers_s3, 1):
    c = ws3.cell(row=3, column=col_idx, value=h)
    c.font = font_header
    c.fill = fill_header
    c.alignment = align_center

sections_data = [
    ("Home", "Hero Brand Sequence", "hero-brand (<HeroBrand />)", "Brand Title & Tagline", "GSAP Title reveal, Viscosity bar"),
    ("Home", "The Vitrine Mosaic", "cm-section (<CollectionsMosaic />)", "9 Category Visual Grid", "Asymmetric card layout"),
    ("Home", "Wave Ribbon", "ribbon (<Ribbon />)", "Animated Ticker", "Breathing SVG text path"),
    ("Home", "Sampler Mobile Slider", "czoom-section is-mobile", "Mobile Piece Gallery", "Touch-driven gallery"),
    ("Home", "Sampler Desktop Zoom", "czoom-section", "Desktop Zoom Transition", "Pinned GSAP scroll zoom"),
    ("Home", "Heritage Video Block", "craft-card-wrapper (#heritage)", "Video Masked Title", "MaskedHeading video canvas"),
    ("Home", "The Turn (Materials)", "tt section (#the-turn)", "Material Properties", "Pinned 5-finish morph"),
    ("Home", "Material Guidance CTA", "mg section (#material-guidance)", "Consultation Callout", "Liquid fill CTAs"),
    ("Home", "Finishes Showcase", "fin section (#finishes)", "Finish Swatch Grid", "Interactive swatch selection"),
    ("Home", "Heritage 5-Step Process", "craft-about-section (#craft)", "Process Narrative + Stats", "5 step items + stat counter cards"),
    ("Home", "Countries Served Loop", "countries-section-wrap (#countries)", "Global Roster", "LogoLoop flag slider"),
    ("Home", "Blogs & Socials", "fs-section (#instagram)", "Media Gallery", "HeroVideoDialog lightboxes"),
    ("Home", "Our Core Philosophy", "phil (#philosophy)", "Brand Principles", "Card reveal grid"),
    ("Home", "Reviews & Testimonials", "bestsellers-section (#reviews)", "Client Feedback", "CardsReveal drag rail"),
    ("Collections", "Collections Grid Wall", "cw (<CollectionsWall />)", "9 Family Cards", "Interactive vitrine hover"),
    ("Collections", "Past 9th Bay CTA", "cw-close", "Bespoke Enquiry Callout", "Specular CTAs"),
    ("Collection Detail", "Masthead Cover", "pl-head pl-head--shot", "Category Hero Image", "Parallax image drift"),
    ("Collection Detail", "Pieces Grid", "pl-grid-wrap", "Product Items Grid", "Velora landscape cards, cart buttons"),
    ("Collection Detail", "Other Families Switcher", "pl-switch", "Category Navigator", "Pill button navigation"),
    ("Catalogue", "Hero Title Banner", "page-hero wrap", "Page Header", "CharCascade title"),
    ("Catalogue", "All Products Grid", "section alt", "Catalog Grid / Empty State", "ProductCard cards"),
    ("Catalogue", "New Arrivals Gallery", "section", "3D Cylinder Showcase", "CircularGallery WebGL/Canvas"),
    ("Product Detail", "Product Hero Split", "page-hero wrap", "Item Specs & CTAs", "Spec list, SpecularButton CTAs"),
    ("Product Detail", "Related Products Rail", "section alt", "Related Items Rail", "DragRail grab & fling"),
    ("Shows", "Hero Title Banner", "page-hero wrap", "Header", "CharCascade title"),
    ("Shows", "Ateliers Gallery", "section alt", "Showroom Plates", "SignatureScroll gallery"),
    ("Shows", "Exhibitions & Events Grid", "section", "Event Schedule", "CardsReveal event cards"),
    ("Shows", "Showroom Visit CTA", "section alt", "Booking CTA", "Direct contact button"),
    ("About", "Hero Title Banner", "page-hero wrap", "Header", "CharCascade headline"),
    ("About", "Heritage Video Canvas", "craft-card-wrapper", "Craft Video Heading", "MaskedHeading video mask"),
    ("About", "Heritage Process Narrative", "craft-about-section", "Narrative & 5 steps", "Process step list"),
    ("About", "Chronology Timeline Rail", "tl (<Timeline />)", "Historical Milestones", "Horizontal drag timeline"),
    ("About", "Core Principles Grid", "section", "Philosophy Cards", "Dilate card reveals"),
    ("Testimonials", "Hero Title Banner", "page-hero wrap", "Header", "CharCascade headline"),
    ("Testimonials", "Review Grid Wall", "section alt", "All Reviews", "CardsReveal quote cards"),
    ("Partners", "Hero Banner & Stats", "page-hero wrap prt-hero", "Header & Roster Stats", "Interactive sector punch tally"),
    ("Partners", "Roster Grid Wall", "prt-wall-sec", "Partner Roster", "Magnetic partner chips with search"),
    ("Partners", "Case Studies Spread", "prt-cases", "3 Enterprise Cases", "50/50 Before-After swipe compare slider"),
    ("Partners", "Client Quotes List", "prt-quotes", "Unprompted Reviews", "Numbered quote cards"),
    ("Partners", "Closing CTA Banner", "prt-cta", "Dark CTA Band", "Partner marquee name ticker"),
    ("Care", "Hero Title Banner", "page-hero wrap", "Header", "CharCascade title"),
    ("Care", "Material Care Tabs", "section alt", "Care Guidance", "Interactive material tab selector"),
    ("Contact", "Hero Title Banner", "page-hero wrap", "Header", "CharCascade title"),
    ("Contact", "Form & Details Split", "section alt", "Enquiry Form", "FluidLabel fields, mailto trigger"),
    ("Contact", "Google Maps Embed", "section", "Moradabad Location", "Interactive coordinates map link"),
    ("FAQ", "Hero Title Banner", "page-hero wrap", "Header", "CharCascade title"),
    ("FAQ", "Inflating Disclosure List", "section alt", "12 Q&A Items", "Accordion expanding pills"),
    ("Legal", "Hero Title Banner", "page-hero wrap", "Header", "CharCascade title"),
    ("Legal", "Document Reader & Index", "section alt", "Terms & Privacy", "Sticky liquid index, paragraph deep-link copy"),
    ("Not Found", "404 Screen", "nf", "Fallback Screen", "Breathing droplet animation"),
    ("Modular/Legacy", "11 Modular Components", "Various (mk2 / components)", "Reusable Component Sections", "Ledger, Measured, CraftIndex, TwoFloors, etc.")
]

row_idx = 4
for s in sections_data:
    ws3.row_dimensions[row_idx].height = 20
    ws3.cell(row=row_idx, column=1, value=s[0]).font = font_bold
    ws3.cell(row=row_idx, column=2, value=s[1]).font = font_regular
    ws3.cell(row=row_idx, column=3, value=s[2]).font = font_regular
    ws3.cell(row=row_idx, column=4, value=s[3]).font = font_regular
    ws3.cell(row=row_idx, column=5, value=s[4]).font = font_regular

    for c in range(1, 6):
        cell = ws3.cell(row=row_idx, column=c)
        cell.border = thin_border
        if row_idx % 2 == 1:
            cell.fill = fill_zebra
        if c in [1]:
            cell.alignment = align_center
        else:
            cell.alignment = align_left
    row_idx += 1


# ---------------------------------------------------------
# SHEET 4: Motion & Micro-interactions
# ---------------------------------------------------------
ws4 = wb.create_sheet(title="Motion & Hovers")

ws4.merge_cells("A1:E1")
ws4["A1"] = "MOTION CATALOGUE & MICRO-INTERACTION AUDIT (14 Hovers / 25 FX)"
ws4["A1"].font = font_title

headers_s4 = ["ID / Code", "Effect Name", "Category", "Technical Implementation", "User Experience Impact"]
ws4.row_dimensions[3].height = 24
for col_idx, h in enumerate(headers_s4, 1):
    c = ws4.cell(row=3, column=col_idx, value=h)
    c.font = font_header
    c.fill = fill_header
    c.alignment = align_center

motion_data = [
    ("E1 / H0", "Chrome Sheen & Specular Highlight", "Micro-interaction", "CSS Custom props --mx/--my set by pointermove rAF", "Cards reflect light like physical brass & copper"),
    ("H1", "Button Magnetism Engine", "Micro-interaction", "Pointer tracking on .btn writing --mag-x/--mag-y", "Buttons pull subtly toward the cursor on proximity"),
    ("H1 (fill)", "Liquid Entry Point Flood", "Micro-interaction", "Radial gradient mask expanding from entry coordinates", "Fluid liquid flood fill effect on hover"),
    ("H1 (press)", "Surge Press Elasticity", "Micro-interaction", "GSAP squash & stretch transform on mouse down", "Tactile feedback when clicking buttons"),
    ("H2", "Asymmetric Slab Radius Warp", "Micro-interaction", "Border-radius morphing dynamically per corner", "Slabs feel malleable and organic"),
    ("H3", "Elastic Underline Navigation", "Micro-interaction", "Liquid pill stretch between nav link targets", "Rubber-band transition between active states"),
    ("H4", "Perimeter Droplet Bead", "Micro-interaction", "CSS motion-path / offset-distance keyframed bead", "Chrome bead runs the perimeter of cards on hover"),
    ("H5", "Dynamic Column Highlight", "Micro-interaction", "12-column grid background tinting on card hover", "Reveals Swiss architectural grid alignment"),
    ("H6", "Segmented Pill Slide", "Micro-interaction", "Lead/trail lag stretch on filter toggles", "Smooth sliding pill selector indicator"),
    ("H7", "Fluid Label & Caret Metal Pool", "Micro-interaction", "Field height stretch & dynamic gradient under caret", "Pill input animation with fluid metallic caret"),
    ("H8", "Dark-Section Pointer Wake", "Micro-interaction", "Trail of low-opacity blurred circles in .deep sections", "Subtle metallic wake following mouse velocity"),
    ("H9", "High-Velocity Strike Sparks", "Micro-interaction", "Pooled DOM node burst of brass filings on click", "Spark particles burst outward on every tap/click"),
    ("H10", "Partner Wall Physics Magnetism", "Micro-interaction", "Inverse square law pointer attraction on roster chips", "Roster chips float toward the cursor"),
    ("H11", "Drag Rail Meniscus Boost", "Micro-interaction", "Dynamic wave amplitude multiplier while dragging", "Gallery rails react to drag force"),
    ("E2", "Meniscus Edge undulation", "Environment FX", "SVG path morphing breathing liquid boundary", "Plates undulate slowly like liquid mercury"),
    ("E4", "Viscosity Top-Edge Scroll Bar", "Environment FX", "Velocity-driven scaleY scroll bar with decimal depth", "Full-width top bar thickening on fast scroll"),
    ("E5", "Droplet Coalesce Preloader", "Environment FX", "SVG goo filter merging 5 droplets into wordmark", "Preloader merges metal droplets into logo"),
    ("E7", "Dilate Scroll Reveal", "Scroll Entrance", "Scale 0.94->1 with radius morph from center", "Elements swell into place rather than sliding"),
    ("E8", "Char Cascade Headlines", "Typography FX", "Staggered center-out vertical character squash", "Headlines squish and settle like liquid metal"),
    ("E9", "Drag Rail Gallery Momentum", "Gallery FX", "GSAP Draggable with momentum, friction & skew", "Grab & fling horizontal product gallery"),
    ("E11", "Pinned Finish Morph Showcase", "Scroll Signature", "ScrollTrigger scrub cross-fading 5 metal finishes", "Object stays while surface material morphs"),
    ("E12", "Wave Ribbon Marquee", "Text FX", "SVG textPath running along undulating breathing paths", "Text rides undulating wave paths"),
    ("E13", "Fluid Fill Numerals", "Stat FX", "Wavy liquid level rising inside outlined figures", "Stat digits fill with liquid metal as you scroll"),
    ("E15", "Viscous Stepper Process", "Timeline FX", "Fat rounded tube filling with fluid, node inflation", "Process stages inflate as fluid level arrives"),
    ("T1", "Meniscus Wipe Transition", "Page Transition", "Molten metal wave overlay flooding page on navigate", "Fluid wave page transition wipe")
]

row_idx = 4
for m in motion_data:
    ws4.row_dimensions[row_idx].height = 20
    ws4.cell(row=row_idx, column=1, value=m[0]).font = font_bold
    ws4.cell(row=row_idx, column=2, value=m[1]).font = font_bold
    ws4.cell(row=row_idx, column=3, value=m[2]).font = font_regular
    ws4.cell(row=row_idx, column=4, value=m[3]).font = font_regular
    ws4.cell(row=row_idx, column=5, value=m[4]).font = font_regular

    for c in range(1, 6):
        cell = ws4.cell(row=row_idx, column=c)
        cell.border = thin_border
        if row_idx % 2 == 1:
            cell.fill = fill_zebra
        if c in [1, 3]:
            cell.alignment = align_center
        else:
            cell.alignment = align_left
    row_idx += 1

# Auto adjust column widths across all sheets
for sheet in wb.worksheets:
    for col in sheet.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if cell.coordinate in sheet.merged_cells:
                continue
            if len(val_str) > max_len:
                max_len = len(val_str)
        sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

wb.save("d:/Code/Taif/TAIF_Website_Metrics_Executive_Summary.xlsx")
print("Excel spreadsheet created successfully!")
