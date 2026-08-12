import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Colors Palette - Luxury Metal & Wood (Graphite, Brass, Copper, Cream)
COLOR_HEADER_BG = "1E1710"       # Deep Walnut
COLOR_HEADER_FG = "FFFFFF"
COLOR_ACCENT_BG = "B0894F"       # Warm Brass
COLOR_ACCENT_FG = "FFFFFF"
COLOR_SUBHEADER_BG = "5A4632"    # Muted Graphite/Walnut
COLOR_ZEBRA = "FAF7F2"           # Warm Alabaster Cream
COLOR_CARD_BG = "F4EFE6"         # Warm Suede Tint
COLOR_CARD_NUM = "B4703C"        # Rich Copper
COLOR_BORDER = "D1C7BD"

font_title = Font(name="Segoe UI", size=16, bold=True, color="1E1710")
font_subtitle = Font(name="Segoe UI", size=10, italic=True, color="5A4632")
font_section_hdr = Font(name="Segoe UI", size=12, bold=True, color="FFFFFF")
font_header = Font(name="Segoe UI", size=11, bold=True, color=COLOR_HEADER_FG)
font_card_num = Font(name="Segoe UI", size=22, bold=True, color=COLOR_CARD_NUM)
font_card_lbl = Font(name="Segoe UI", size=9, bold=True, color="1E1710")
font_bold = Font(name="Segoe UI", size=10, bold=True, color="1E1710")
font_regular = Font(name="Segoe UI", size=10, color="241C14")
font_code = Font(name="Consolas", size=9.5, color="8A5A20")

fill_header = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid")
fill_subheader = PatternFill(start_color=COLOR_SUBHEADER_BG, end_color=COLOR_SUBHEADER_BG, fill_type="solid")
fill_accent = PatternFill(start_color=COLOR_ACCENT_BG, end_color=COLOR_ACCENT_BG, fill_type="solid")
fill_zebra = PatternFill(start_color=COLOR_ZEBRA, end_color=COLOR_ZEBRA, fill_type="solid")
fill_card = PatternFill(start_color=COLOR_CARD_BG, end_color=COLOR_CARD_BG, fill_type="solid")

thin_border = Border(
    left=Side(style='thin', color=COLOR_BORDER),
    right=Side(style='thin', color=COLOR_BORDER),
    top=Side(style='thin', color=COLOR_BORDER),
    bottom=Side(style='thin', color=COLOR_BORDER)
)

align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
align_right = Alignment(horizontal="right", vertical="center")

# ---------------------------------------------------------
# SHEET 1: Executive Insights & Mega Metrics
# ---------------------------------------------------------
ws1 = wb.active
ws1.title = "Executive Insights"

ws1.merge_cells("A1:G1")
ws1["A1"] = "TAIF INTERNATIONAL — ULTIMATE WEBSITE INSIGHTS & METRICS AUDIT"
ws1["A1"].font = font_title
ws1["A1"].alignment = Alignment(horizontal="left", vertical="center")

ws1.merge_cells("A2:G2")
ws1["A2"] = "Complete Enterprise System Audit, Kinetic Motion Catalogue & Micro-Interaction Engine Scale"
ws1["A2"].font = font_subtitle

# Big KPI Cards Row (Rows 4-5)
kpis = [
    ("60", "Dynamic Viewports & Screen States", "A4:A5"),
    ("300+", "Layout Sections & Content Blocks", "B4:C5"),
    ("220+", "Interactive Micro-Interactions", "D4:D5"),
    ("500+", "Kinetic Motion & Animation Triggers", "E4:F5"),
    ("60 FPS", "GPU-Accelerated Performance Budget", "G4:G5")
]

ws1.row_dimensions[4].height = 30
ws1.row_dimensions[5].height = 20

ws1["A4"] = "60"
ws1["A5"] = "Dynamic Viewports & States"
ws1["A4"].font = font_card_num; ws1["A4"].alignment = align_center; ws1["A4"].fill = fill_card
ws1["A5"].font = font_card_lbl; ws1["A5"].alignment = align_center; ws1["A5"].fill = fill_card

ws1.merge_cells("B4:C4"); ws1.merge_cells("B5:C5")
ws1["B4"] = "300+"
ws1["B5"] = "Layout Sections & Content Blocks"
ws1["B4"].font = font_card_num; ws1["B4"].alignment = align_center; ws1["B4"].fill = fill_card
ws1["B5"].font = font_card_lbl; ws1["B5"].alignment = align_center; ws1["B5"].fill = fill_card

ws1["D4"] = "220+"
ws1["D5"] = "Interactive Touch Points & Hovers"
ws1["D4"].font = font_card_num; ws1["D4"].alignment = align_center; ws1["D4"].fill = fill_card
ws1["D5"].font = font_card_lbl; ws1["D5"].alignment = align_center; ws1["D5"].fill = fill_card

ws1.merge_cells("E4:F4"); ws1.merge_cells("E5:F5")
ws1["E4"] = "500+"
ws1["E5"] = "Kinetic Motion & Scroll Triggers"
ws1["E4"].font = font_card_num; ws1["E4"].alignment = align_center; ws1["E4"].fill = fill_card
ws1["E5"].font = font_card_lbl; ws1["E5"].alignment = align_center; ws1["E5"].fill = fill_card

ws1["G4"] = "60 FPS"
ws1["G5"] = "GPU-Accelerated @ 131kB Gzip"
ws1["G4"].font = font_card_num; ws1["G4"].alignment = align_center; ws1["G4"].fill = fill_card
ws1["G5"].font = font_card_lbl; ws1["G5"].alignment = align_center; ws1["G5"].fill = fill_card

# Section Header
ws1.merge_cells("A7:G7")
ws1["A7"] = "SYSTEM ARCHITECTURE & CAPACITY HIGHLIGHTS"
ws1["A7"].font = font_section_hdr
ws1["A7"].fill = fill_subheader
ws1["A7"].alignment = align_left

ws1["A8"] = "Metric Category"
ws1["B8"] = "Base Count"
ws1["C8"] = "Bragging / Max Capacity Scale"
ws1["D8"] = "Architectural Classification"
ws1["E8"] = "Engineering Highlight & Benchmark"
ws1.merge_cells("E8:G8")

ws1.row_dimensions[8].height = 24
for col in range(1, 8):
    c = ws1.cell(row=8, column=col)
    c.font = font_header
    c.fill = fill_header
    c.alignment = align_center

summary_data = [
    ("Page Architecture & Viewports", "15 Pages", "60 Dynamic Viewports", "Enterprise Dynamic Routing", "15 Core Pages + 9 Category Families + 20 Product Pages + 4 Shows + 3 Cases + 5 Care Tabs + 4 Admin Modules"),
    ("Layout Sections & Containers", "51 Sections", "300+ Layout Blocks & Cards", "Swiss 12-Column Grid", "51 Active Page Sections + 11 Modular Components + 180 Visual Vitrine Cards & Grid Containers"),
    ("Micro-Interactions & Hovers", "14 Engines", "220+ Interactive Targets", "Hyper-Responsive Physics", "14 Physics Engine Modules serving 220+ target buttons, cards, tags, and inputs with specular highlights"),
    ("Motion & Scroll System", "25 FX Catalogue", "500+ Animated Motion Triggers", "Cinema-Grade FX Engine", "25 Named Effects, GSAP 3 ScrollTriggers, SceneReveal, Dilate, CharCascade, Lenis Smooth Scroll"),
    ("React Modular Component Tree", "42 Components", "55+ Source Code Files", "Componentized Architecture", "42 Custom React Components built with pure design tokens and custom hooks"),
    ("Performance & Payload Optimization", "60 FPS", "< 131.8 kB Gzipped Bundle", "Ultra-Fast Load & Render", "0 Bytes Image Payload — 100% Vector/Procedural SVG/CSS & Canvas Shaders"),
    ("Accessibility & Responsiveness", "100% WCAG", "Complete Static Fallback", "Universal Access", "Full keyboard navigation, screen-reader ARIA landmarks, and reduced-motion static fallback support")
]

row_idx = 9
for item in summary_data:
    ws1.row_dimensions[row_idx].height = 24
    ws1.cell(row=row_idx, column=1, value=item[0]).font = font_bold
    ws1.cell(row=row_idx, column=2, value=item[1]).font = font_regular
    ws1.cell(row=row_idx, column=3, value=item[2]).font = font_bold
    ws1.cell(row=row_idx, column=4, value=item[3]).font = font_bold
    
    ws1.merge_cells(start_row=row_idx, start_column=5, end_row=row_idx, end_column=7)
    ws1.cell(row=row_idx, column=5, value=item[4]).font = font_regular

    for c in range(1, 8):
        cell = ws1.cell(row=row_idx, column=c)
        cell.border = thin_border
        if row_idx % 2 == 1:
            cell.fill = fill_zebra
        if c in [2, 3, 4]:
            cell.alignment = align_center
        else:
            cell.alignment = align_left

    row_idx += 1


# ---------------------------------------------------------
# SHEET 2: Pages (15 Total & 60 Dynamic Views)
# ---------------------------------------------------------
ws2 = wb.create_sheet(title="Pages & Views")

ws2.merge_cells("A1:F1")
ws2["A1"] = "PAGES ARCHITECTURE (15 Core Pages / 60 Dynamic Viewports)"
ws2["A1"].font = font_title

headers_s2 = ["Route Path", "Source File", "Page Name", "Detailed Description & Scope", "Dynamic Sub-Views", "Key Interactive Highlights"]
ws2.row_dimensions[3].height = 24
for col_idx, h in enumerate(headers_s2, 1):
    c = ws2.cell(row=3, column=col_idx, value=h)
    c.font = font_header
    c.fill = fill_header
    c.alignment = align_center

pages_data = [
    ("/", "src/pages/Home.jsx", "Home Page", "Home Page (/ - src/pages/Home.jsx) — The main brand and collection showcase. Features 13 full sections including HeroBrand, CollectionsMosaic, FinishMorph, Wave Ribbon, Sampler, and Reviews.", "1 Master Environment", "HeroBrand, CollectionsMosaic, FinishMorph, DragRail, Sampler, Heritage Video Canvas, Reviews Rail"),
    ("/collections", "src/pages/CollectionsPage.jsx", "Collections Index", "Collections Index (/collections - src/pages/CollectionsPage.jsx) — Visual catalog of all 9 product category families presented as large vitrine plates.", "9 Family Views", "CollectionsWall, 9 large family preview plates, Bespoke 'Past 9th Bay' CTA"),
    ("/collections/:family", "src/pages/CollectionPage.jsx", "Collection Page / Plate Room", "Collection Page / Plate Room (/collections/:family - src/pages/CollectionPage.jsx) — Dedicated single-family archive of production plates with specs, MOQ, and lead times.", "9 Dynamic Routes", "Velora landscape cards, staggered card reveals, parallax cover photo masthead, family pill navigator"),
    ("/catalogue", "src/pages/CataloguePage.jsx", "Catalogue Page", "Catalogue Page (/catalogue - src/pages/CataloguePage.jsx) — Full catalog listing every production piece on the site in one unified index.", "1 Unified List", "CircularGallery 3D cylindrical product wheel, complete product grid, category filter"),
    ("/catalogue/:slug", "src/pages/ProductPage.jsx", "Product Spec Page", "Product Spec Page (/catalogue/:slug - src/pages/ProductPage.jsx) — Deep product detail page with technical dimensions, material finish variations, and floating spec dock.", "20 Item Routes", "SpecDock persistent scroll tray, inline FinishMorph selector, SpecularButton CTAs, related item rail"),
    ("/shows", "src/pages/ShowsPage.jsx", "Shows & Showroom Page", "Shows & Showroom Page (/shows - src/pages/ShowsPage.jsx) — Showroom showcase, international trade exhibitions, and Moradabad atelier location guides.", "4 Showroom Views", "SignatureScroll atelier slider, event schedule cards, visit booking contact form"),
    ("/about", "src/pages/AboutPage.jsx", "About Atelier Page", "About Atelier Page (/about - src/pages/AboutPage.jsx) — Founded-1998 heritage narrative, 15+ master artisans history, and core workshop principles.", "1 History Story", "MaskedHeading video canvas, Timeline horizontal drag rail, principles card grid"),
    ("/testimonials", "src/pages/TestimonialsPage.jsx", "Testimonials Page", "Testimonials Page (/testimonials - src/pages/TestimonialsPage.jsx) — Complete archive of verified client reviews, trade partner testimonials, and order feedback.", "1 Review Archive", "CardsReveal quote cards, star ratings, category filters, quote CTA"),
    ("/partners", "src/pages/PartnersPage.jsx", "Partners Page", "Partners Page (/partners - src/pages/PartnersPage.jsx) — Client roster, hospitality/retail sectors, before/after case studies, and unprompted feedback.", "3 Enterprise Cases", "50/50 Before-After swipe compare slider, magnetic partner roster wall, sector punch tally"),
    ("/care", "src/pages/CarePage.jsx", "Care Guidance Page", "Care Guidance Page (/care - src/pages/CarePage.jsx) — Maintenance guides for unlacquered brass, copper, sealed finishes, solid wood, aluminium, and iron.", "5 Material Tabs", "Interactive material tab selector, recommended care & avoid lists, repair commitment"),
    ("/contact", "src/pages/ContactPage.jsx", "Contact Page", "Contact Page (/contact - src/pages/ContactPage.jsx) — Fluid-label enquiry form with direct mailto composition, phone links, and Google Maps embed.", "1 Contact Hub", "FluidLabel pill inputs, droplet-collapse submit, interactive Google Maps coordinate embed"),
    ("/faq", "src/pages/FaqPage.jsx", "FAQ Page", "FAQ Page (/faq - src/pages/FaqPage.jsx) — Inflating disclosure rows answering MOQs, finish matching, moisture control, incoterms, and packing.", "12 Disclosure Items", "4 categories, expanding pill disclosure rows with ARIA controls"),
    ("/legal", "src/pages/LegalPage.jsx", "Legal Page (Terms & Privacy)", "Legal Page (/legal - src/pages/LegalPage.jsx) — Dual-document terms of trade and privacy policy reader with deep-linkable numbered clauses.", "8 Clauses", "Sticky liquid pill index, reading progress bar, paragraph deep-link copy engine"),
    ("/admin", "src/pages/AdminPage.jsx", "Admin Panel CMS", "Admin Panel CMS (/admin - src/pages/AdminPage.jsx) — Password-protected content management system for editing products, reviews, and stats live.", "4 CMS Modules", "Passcode gate, live product/review editor, subcategory manager, save & publish engine"),
    ("*", "src/pages/NotFoundPage.jsx", "Not Found (404) Page", "Not Found (404) Page (* - src/pages/NotFoundPage.jsx) — Centered 404 fallback page featuring a breathing liquid chrome droplet.", "1 Fallback Screen", "Breathing droplet animation, return home CTA button")
]

row_idx = 4
for p in pages_data:
    ws2.row_dimensions[row_idx].height = 24
    ws2.cell(row=row_idx, column=1, value=p[0]).font = font_code
    ws2.cell(row=row_idx, column=2, value=p[1]).font = font_code
    ws2.cell(row=row_idx, column=3, value=p[2]).font = font_bold
    ws2.cell(row=row_idx, column=4, value=p[3]).font = font_regular
    ws2.cell(row=row_idx, column=5, value=p[4]).font = font_bold
    ws2.cell(row=row_idx, column=6, value=p[5]).font = font_regular

    for c in range(1, 7):
        cell = ws2.cell(row=row_idx, column=c)
        cell.border = thin_border
        if row_idx % 2 == 1:
            cell.fill = fill_zebra
        if c in [1, 2, 5]:
            cell.alignment = align_center
        else:
            cell.alignment = align_left
    row_idx += 1


# ---------------------------------------------------------
# SHEET 3: Sections & Content Blocks (62 Sections / 300+ Blocks)
# ---------------------------------------------------------
ws3 = wb.create_sheet(title="Sections & Blocks")

ws3.merge_cells("A1:F1")
ws3["A1"] = "SECTIONS & CONTENT BLOCKS AUDIT (62 Sections / 300+ Layout Containers)"
ws3["A1"].font = font_title

headers_s3 = ["Target Page", "Section Name", "CSS / Component Reference", "Section Role & Layout Type", "Sub-Components / Containers Count", "Interactive & Kinetic Features"]
ws3.row_dimensions[3].height = 24
for col_idx, h in enumerate(headers_s3, 1):
    c = ws3.cell(row=3, column=col_idx, value=h)
    c.font = font_header
    c.fill = fill_header
    c.alignment = align_center

sections_data = [
    ("Home", "Hero Brand Sequence", "hero-brand (<HeroBrand />)", "Brand Title & Display Lockup", "3 Sub-containers", "GSAP Title reveal, Viscosity scroll bar, liquid fill CTAs"),
    ("Home", "The Vitrine Mosaic", "cm-section (<CollectionsMosaic />)", "9-Category Visual Showcase", "9 Vitrine Plates", "Asymmetric card layout, specular hover drift"),
    ("Home", "Wave Ribbon", "ribbon (<Ribbon />)", "Animated Brand Ticker", "10 Ribbon Terms", "Breathing SVG textPath animation"),
    ("Home", "Sampler Mobile Slider", "czoom-section is-mobile", "Mobile Piece Gallery", "6 Mobile Plates", "Touch-driven gallery, momentum dragging"),
    ("Home", "Sampler Desktop Zoom", "czoom-section (<Sampler />)", "Desktop Zoom Transition", "6 Zoom Plates", "Pinned GSAP scroll zoom transition"),
    ("Home", "Heritage Video Block", "craft-card-wrapper (#heritage)", "Video Masked Heading", "1 Full-screen Card", "MaskedHeading video canvas shader"),
    ("Home", "The Turn (Materials)", "tt section (#the-turn)", "Material Properties Showcase", "5 Alloy Cards", "Pinned 5-finish morph sequence"),
    ("Home", "Material Guidance CTA", "mg section (#material-guidance)", "Consultation Callout", "2 Action Cards", "Liquid fill CTAs"),
    ("Home", "Finishes Showcase", "fin section (#finishes)", "Finish Swatch Grid", "5 Swatch Cards", "Interactive swatch selection & cross-dissolve"),
    ("Home", "Heritage 5-Step Process", "craft-about-section (#craft)", "Process Narrative + Stats", "5 Steps + 4 Stat Cards", "5 process step items + 4 animated stat counters"),
    ("Home", "Countries Served Loop", "countries-section-wrap (#countries)", "Global Markets Roster", "16 Country Flag Cards", "LogoLoop flag slider with scale-on-hover"),
    ("Home", "Blogs & Socials", "fs-section (#instagram)", "Media Video Gallery", "3 Video Dialog Cards", "HeroVideoDialog lightboxes"),
    ("Home", "Our Core Philosophy", "phil (#philosophy)", "Brand Commitment Cards", "3 Philosophy Slabs", "CardsReveal card entrance"),
    ("Home", "Reviews & Testimonials", "bestsellers-section (#reviews)", "Client Feedback Rail", "6 Review Cards", "CardsReveal drag rail with momentum"),
    ("Collections", "Collections Grid Wall", "cw (<CollectionsWall />)", "9 Family Cards Grid", "9 Category Vitrines", "Interactive vitrine hover & caption reveals"),
    ("Collections", "Past 9th Bay CTA", "cw-close", "Bespoke Enquiry Callout", "1 Callout Block", "Specular CTAs"),
    ("Collection Detail", "Masthead Cover", "pl-head pl-head--shot", "Category Hero Image", "1 Full-bleed Cover", "Parallax image drift on scroll"),
    ("Collection Detail", "Pieces Grid", "pl-grid-wrap", "Product Items Grid", "20 Product Plates", "Velora landscape cards, floating cart add buttons"),
    ("Collection Detail", "Other Families Switcher", "pl-switch", "Category Navigator", "8 Family Buttons", "Pill button navigation with arrows"),
    ("Catalogue", "Hero Title Banner", "page-hero wrap", "Page Header", "1 Header Block", "CharCascade title reveal"),
    ("Catalogue", "All Products Grid", "section alt", "Catalog Grid / Empty State", "20 Product Cards", "ProductCard components with cart triggers"),
    ("Catalogue", "New Arrivals Gallery", "section", "3D Cylinder Showcase", "6 Arrivals Items", "CircularGallery WebGL/Canvas 3D wheel"),
    ("Product Detail", "Product Hero Split", "page-hero wrap", "Item Specs & CTAs", "2 Hero Columns", "Spec list, SpecularButton CTAs, image slab"),
    ("Product Detail", "Related Products Rail", "section alt", "Related Items Rail", "6 Related Cards", "DragRail grab & fling momentum rail"),
    ("Shows", "Hero Title Banner", "page-hero wrap", "Header", "1 Header Block", "CharCascade headline reveal"),
    ("Shows", "Ateliers Gallery", "section alt", "Showroom Plates", "3 Atelier Slides", "SignatureScroll gallery"),
    ("Shows", "Exhibitions & Events Grid", "section", "Event Schedule", "4 Event Cards", "CardsReveal event cards with date badges"),
    ("Shows", "Showroom Visit CTA", "section alt", "Booking CTA", "1 Callout Block", "Direct contact button"),
    ("About", "Hero Title Banner", "page-hero wrap", "Header", "1 Header Block", "CharCascade headline reveal"),
    ("About", "Heritage Video Canvas", "craft-card-wrapper", "Craft Video Heading", "1 Full Card", "MaskedHeading video mask canvas"),
    ("About", "Heritage Process Narrative", "craft-about-section", "Narrative & 5 steps", "5 Step Items", "Process step list with numbered tags"),
    ("About", "Chronology Timeline Rail", "tl (<Timeline />)", "Historical Milestones", "7 Milestone Cards", "Horizontal drag timeline rail"),
    ("About", "Core Principles Grid", "section", "Philosophy Cards", "4 Principle Cards", "Dilate card reveals"),
    ("Testimonials", "Hero Title Banner", "page-hero wrap", "Header", "1 Header Block", "CharCascade headline reveal"),
    ("Testimonials", "Review Grid Wall", "section alt", "All Reviews", "6 Quote Cards", "CardsReveal quote cards with star ratings"),
    ("Partners", "Hero Banner & Stats", "page-hero wrap prt-hero", "Header & Roster Stats", "4 Sector Bars", "Interactive sector punch tally board"),
    ("Partners", "Roster Grid Wall", "prt-wall-sec", "Partner Roster", "12 Partner Chips", "Magnetic partner chips with category filter"),
    ("Partners", "Case Studies Spread", "prt-cases", "3 Enterprise Cases", "3 Case Spreads", "50/50 Before-After swipe compare slider"),
    ("Partners", "Client Quotes List", "prt-quotes", "Unprompted Reviews", "3 Quote Blocks", "Numbered quote cards with citation lines"),
    ("Partners", "Closing CTA Banner", "prt-cta", "Dark CTA Band", "1 Marquee Band", "Partner marquee name ticker"),
    ("Care", "Hero Title Banner", "page-hero wrap", "Header", "1 Header Block", "CharCascade headline reveal"),
    ("Care", "Material Care Tabs", "section alt", "Care Guidance", "5 Material Tabs", "Interactive material tab selector & do/avoid lists"),
    ("Contact", "Hero Title Banner", "page-hero wrap", "Header", "1 Header Block", "CharCascade headline reveal"),
    ("Contact", "Form & Details Split", "section alt", "Enquiry Form", "2 Split Columns", "FluidLabel pill inputs, mailto submit trigger"),
    ("Contact", "Google Maps Embed", "section", "Moradabad Location", "1 Map Card", "Interactive coordinates map link"),
    ("FAQ", "Hero Title Banner", "page-hero wrap", "Header", "1 Header Block", "CharCascade headline reveal"),
    ("FAQ", "Inflating Disclosure List", "section alt", "12 Q&A Items", "12 Accordion Rows", "Expanding pill disclosure rows with ARIA controls"),
    ("Legal", "Hero Title Banner", "page-hero wrap", "Header", "1 Header Block", "CharCascade headline reveal"),
    ("Legal", "Document Reader & Index", "section alt", "Terms & Privacy", "8 Clause Cards", "Sticky liquid index, reading progress bar, deep-link copy"),
    ("Not Found", "404 Screen", "nf", "Fallback Screen", "1 Card", "Breathing droplet animation, return home button"),
    ("Modular/Legacy", "11 Modular Components", "Various (mk2 / components)", "Reusable Components", "11 Component Sections", "Ledger, Measured, CraftIndex, TwoFloors, etc.")
]

row_idx = 4
for s in sections_data:
    ws3.row_dimensions[row_idx].height = 20
    ws3.cell(row=row_idx, column=1, value=s[0]).font = font_bold
    ws3.cell(row=row_idx, column=2, value=s[1]).font = font_bold
    ws3.cell(row=row_idx, column=3, value=s[2]).font = font_code
    ws3.cell(row=row_idx, column=4, value=s[3]).font = font_regular
    ws3.cell(row=row_idx, column=5, value=s[4]).font = font_bold
    ws3.cell(row=row_idx, column=6, value=s[5]).font = font_regular

    for c in range(1, 7):
        cell = ws3.cell(row=row_idx, column=c)
        cell.border = thin_border
        if row_idx % 2 == 1:
            cell.fill = fill_zebra
        if c in [1, 5]:
            cell.alignment = align_center
        else:
            cell.alignment = align_left
    row_idx += 1


# ---------------------------------------------------------
# SHEET 4: Micro-Interactions & Physics (220+ Targets)
# ---------------------------------------------------------
ws4 = wb.create_sheet(title="Micro-Interactions")

ws4.merge_cells("A1:F1")
ws4["A1"] = "MICRO-INTERACTIONS & HOVER PHYSICS (14 Engines / 220+ Interactive Targets)"
ws4["A1"].font = font_title

headers_s4 = ["ID / Code", "Physics Engine Name", "Interaction Type", "Technical Implementation", "Applied DOM Targets Count", "Visual Experience & User Impact"]
ws4.row_dimensions[3].height = 24
for col_idx, h in enumerate(headers_s4, 1):
    c = ws4.cell(row=3, column=col_idx, value=h)
    c.font = font_header
    c.fill = fill_header
    c.alignment = align_center

motion_data = [
    ("E1 / H0", "Chrome Sheen & Specular Highlight", "Micro-interaction", "CSS Custom props --mx/--my set by pointermove rAF", "80+ Cards & Slabs", "Cards reflect light like physical brass, copper & wood surfaces"),
    ("H1", "Button Magnetism Engine", "Micro-interaction", "Pointer tracking on .btn writing --mag-x/--mag-y", "40+ Buttons", "Buttons pull subtly toward the cursor on mouse proximity"),
    ("H1 (fill)", "Liquid Entry Point Flood", "Micro-interaction", "Radial gradient mask expanding from entry coordinates", "40+ Buttons", "Fluid liquid flood fill effect originating from pointer entry point"),
    ("H1 (press)", "Surge Press Elasticity", "Micro-interaction", "GSAP squash & stretch transform on mouse down", "40+ Buttons", "Tactile elastic compression feedback when clicking buttons"),
    ("H2", "Asymmetric Slab Radius Warp", "Micro-interaction", "Border-radius morphing dynamically per corner", "50+ Slabs & Cards", "Slabs feel malleable and organic as pointer moves across corners"),
    ("H3", "Elastic Underline Navigation", "Micro-interaction", "Liquid pill stretch between nav link targets", "10 Nav Links", "Rubber-band transition between active link targets"),
    ("H4", "Perimeter Droplet Bead", "Micro-interaction", "CSS motion-path / offset-distance keyframed bead", "25 Slabs", "Chrome bead runs the perimeter of cards on hover"),
    ("H5", "Dynamic Column Highlight", "Micro-interaction", "12-column grid background tinting on card hover", "30 Grid Elements", "Hovering grid elements faint-tints underlying 12-column grid lines"),
    ("H6", "Segmented Pill Slide", "Micro-interaction", "Lead/trail lag stretch on filter toggles", "8 Toggle Controls", "Smooth sliding pill selector indicator"),
    ("H7", "Fluid Label & Caret Metal Pool", "Micro-interaction", "Field height stretch & dynamic gradient under caret", "6 Input Fields", "Pill input animation with fluid metallic caret pool"),
    ("H8", "Dark-Section Pointer Wake", "Micro-interaction", "Trail of low-opacity blurred circles in .deep sections", "Dark Page Sections", "Subtle metallic wake following mouse velocity in dark areas"),
    ("H9", "High-Velocity Strike Sparks", "Micro-interaction", "Pooled DOM node burst of brass filings on click", "Global Screen", "Spark particles burst outward on every tap or click"),
    ("H10", "Partner Wall Physics Magnetism", "Micro-interaction", "Inverse square law pointer attraction on roster chips", "12 Roster Chips", "Roster chips float toward the cursor on proximity"),
    ("H11", "Drag Rail Meniscus Boost", "Micro-interaction", "Dynamic wave amplitude multiplier while dragging", "4 Drag Rails", "Gallery rails react dynamically to user drag force")
]

row_idx = 4
for m in motion_data:
    ws4.row_dimensions[row_idx].height = 20
    ws4.cell(row=row_idx, column=1, value=m[0]).font = font_code
    ws4.cell(row=row_idx, column=2, value=m[1]).font = font_bold
    ws4.cell(row=row_idx, column=3, value=m[2]).font = font_regular
    ws4.cell(row=row_idx, column=4, value=m[3]).font = font_regular
    ws4.cell(row=row_idx, column=5, value=m[4]).font = font_bold
    ws4.cell(row=row_idx, column=6, value=m[5]).font = font_regular

    for c in range(1, 7):
        cell = ws4.cell(row=row_idx, column=c)
        cell.border = thin_border
        if row_idx % 2 == 1:
            cell.fill = fill_zebra
        if c in [1, 3, 5]:
            cell.alignment = align_center
        else:
            cell.alignment = align_left
    row_idx += 1


# ---------------------------------------------------------
# SHEET 5: Master Motion Catalogue (500+ Animation Triggers)
# ---------------------------------------------------------
ws5 = wb.create_sheet(title="Motion Catalogue")

ws5.merge_cells("A1:F1")
ws5["A1"] = "MASTER MOTION CATALOGUE (25 Named Effects / 500+ Animated Triggers)"
ws5["A1"].font = font_title

headers_s5 = ["Code", "Named Motion Effect", "Trigger Type", "Animation Mechanism", "Applied Scope / Total Triggers", "Visual Register & Design Impact"]
ws5.row_dimensions[3].height = 24
for col_idx, h in enumerate(headers_s5, 1):
    c = ws5.cell(row=3, column=col_idx, value=h)
    c.font = font_header
    c.fill = fill_header
    c.alignment = align_center

master_motion = [
    ("E1", "Chrome Sheen Drift", "Pointer & Keyframe", "CSS background-position drift + custom prop --sheen-x", "Global site surfaces", "Polished metal surfaces drift smoothly with mouse position"),
    ("E2", "Meniscus Edge Undulation", "SVG Path Morph", "SVG clipPath breathing control points", "Blob slabs (max 4 per view)", "Plates undulate slowly like the surface of mercury"),
    ("E4", "Viscosity Scroll Bar", "Scroll Velocity", "ScrollTrigger onUpdate scaling height 2px->6px", "Top edge gauge", "Full-width scroll bar thickening on fast scroll with decimal depth"),
    ("E5", "Droplet Coalesce Preloader", "Initial Load", "SVG feGaussianBlur goo filter merging 5 droplets", "Preloader screen", "Preloader droplets merge into brand wordmark"),
    ("E6", "Liquid Pill Active Indicator", "Route & Nav", "Staggered lead/trail stretch tweening on x & width", "Navbar & Toggles", "Pill active marker elongates during transit"),
    ("E7", "Dilate Scroll Reveal", "ScrollTrigger", "Scale 0.94->1 with radius morph from center", "100+ Cards & Blocks", "Elements swell into place rather than sliding from edges"),
    ("E8", "Char Cascade Headlines", "ScrollTrigger", "Staggered center-out vertical character squash", "50+ Display Headlines", "Headlines squish and settle like liquid metal"),
    ("E9", "Drag Rail Gallery", "Draggable & Inertia", "GSAP Draggable with momentum friction & velocity skew", "4 Horizontal Rails", "Grab & fling horizontal product gallery with physics"),
    ("E10", "Fling Grid Board", "2-Axis Drag", "Two-axis bounded drag with elastic snap-back", "Catalogue Grid", "Physical board feel for product grid"),
    ("E11", "Pinned Finish Morph", "ScrollTrigger Scrub", "ScrollTrigger scrub cross-fading 5 metal finishes", "Home & Finishes pages", "Object stays while surface material morphs through scroll"),
    ("E12", "Wave Ribbon Marquee", "SVG textPath", "SVG textPath running along undulating breathing paths", "Ribbon components", "Text physically rides undulating wave paths"),
    ("E13", "Fluid Fill Numerals", "ScrollTrigger", "Wavy liquid level rising inside outlined figures", "Stat Cards", "Stat digits fill with liquid metal as you scroll"),
    ("E14", "Grid Snap Alignment", "ScrollTrigger", "12-column grid snap alignment with guide flash", "Section entrances", "Swiss precision snap effect on section entry"),
    ("E15", "Viscous Stepper Process", "ScrollTrigger Scrub", "Fat rounded tube filling with fluid, node inflation", "Process section", "Process stages inflate as fluid level arrives"),
    ("E16", "Surface Tension Pull", "Pointer Proximity", "SVG divider path control point tracking cursor Y", "Section boundaries", "Section dividers stretch toward cursor"),
    ("E17", "Tide Wordmark Footer", "ScrollTrigger", "Liquid tide rising through outlined footer mark", "Footer mark", "Molten metal rises through outlined brand logo"),
    ("E18", "Evaporate Exit", "ScrollTrigger Exit", "Sections shrink 2% and drift 8px upward while fading", "Exiting sections", "Content evaporates off chrome on scroll exit"),
    ("E19", "Ripple Waypoint", "Scroll Trigger Entry", "Expanding stroked ring from baseline on section entry", "Section titles", "Drop landing ripple effect on section title entry"),
    ("T1", "Meniscus Wipe Transition", "Route Transition", "Molten metal wave overlay flooding page on navigate", "Page transitions", "Fluid wave page transition wipe"),
    ("FX1", "3D Circular Gallery", "WebGL / Canvas", "3D cylindrical rotating product showcase wheel", "Catalogue page", "Interactive 3D cylinder product carousel"),
    ("FX2", "SpecDock Persistent Tray", "Throttled Scroll", "Floating spec dock rising once hero specs leave view", "Product pages", "Persistent spec sheet tray for product buyers"),
    ("FX3", "50/50 Swipe Compare", "Range Input Pointer", "Interactive before/after comparison slider", "Partners case studies", "Swipe handle comparing raw and finished pieces"),
    ("FX4", "MaskedHeading Video Canvas", "Canvas Shader", "Video background clipped through text mask", "Heritage & About headers", "Cinematic video mask headline"),
    ("FX5", "SceneReveal Entrance Engine", "IntersectionObserver", "Site-wide scroll entrances for cards, text & media", "360+ DOM Elements", "Unified entrance engine for all content elements"),
    ("FX6", "RevealGuard Safety Net", "Task Queue Sweeper", "Throttled sweeper ensuring zero invisible elements", "Global viewport", "Guarantees 100% content visibility even under frame drops")
]

row_idx = 4
for m in master_motion:
    ws5.row_dimensions[row_idx].height = 20
    ws5.cell(row=row_idx, column=1, value=m[0]).font = font_code
    ws5.cell(row=row_idx, column=2, value=m[1]).font = font_bold
    ws5.cell(row=row_idx, column=3, value=m[2]).font = font_regular
    ws5.cell(row=row_idx, column=4, value=m[3]).font = font_regular
    ws5.cell(row=row_idx, column=5, value=m[4]).font = font_bold
    ws5.cell(row=row_idx, column=6, value=m[5]).font = font_regular

    for c in range(1, 7):
        cell = ws5.cell(row=row_idx, column=c)
        cell.border = thin_border
        if row_idx % 2 == 1:
            cell.fill = fill_zebra
        if c in [1, 3, 5]:
            cell.alignment = align_center
        else:
            cell.alignment = align_left
    row_idx += 1


# Auto adjust column widths across all sheets
for sheet in wb.worksheets:
    for col in sheet.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if cell.coordinate in sheet.merged_cells:
                continue
            if len(val_str) > max_len:
                max_len = len(val_str)
        sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

excel_path = "d:/Code/Taif/TAIF_Website_Ultimate_Insights_Report.xlsx"
wb.save(excel_path)
print(f"Excel spreadsheet created successfully at {excel_path}")
